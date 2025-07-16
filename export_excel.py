"""
Excel Export Module

Handles exporting test cases to Excel format using pandas.
"""

from typing import List, Dict, Any
import pandas as pd
from pathlib import Path


def save_testcases_to_excel(testcases: List[Dict[str, Any]], output_path: str) -> None:
    """
    Save test cases to an Excel file.
    
    Args:
        testcases: List of test case dictionaries
        output_path: Path to the output Excel file
    """
    
    if not testcases:
        raise ValueError("No test cases to export")
    
    # Prepare data for DataFrame
    rows = []
    for tc in testcases:
        # Convert steps list to numbered string
        steps_text = ""
        if isinstance(tc.get('steps'), list):
            steps_text = "\n".join([f"{i+1}. {step}" for i, step in enumerate(tc['steps'])])
        else:
            steps_text = str(tc.get('steps', ''))
        
        row = {
            'Test ID': tc.get('test_id', ''),
            'Feature': tc.get('feature', ''),
            'Test Case Title': tc.get('title', ''),
            'Test Steps': steps_text,
            'Expected Result': tc.get('expected_result', ''),
            'Priority': tc.get('priority', 'Medium'),
            'Status': 'Not Executed',  # Default status
            'Actual Result': '',  # Empty for manual filling
            'Notes': ''  # Empty for manual notes
        }
        rows.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Create Excel writer with formatting
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Test Cases', index=False)
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Test Cases']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set reasonable limits for column width
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Set row height for better readability
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 20
        
        # Make headers bold
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set text wrapping for test steps and expected results
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column in [4, 5]:  # Test Steps and Expected Result columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    print(f"✅ Excel file saved successfully: {output_path}")


def create_sample_testcases() -> List[Dict[str, Any]]:
    """Create sample test cases for testing purposes."""
    return [
        {
            'feature': 'User Authentication',
            'test_id': 'TC001',
            'title': 'Verify successful login with valid credentials',
            'steps': [
                'Navigate to login page',
                'Enter valid username',
                'Enter valid password',
                'Click Login button'
            ],
            'expected_result': 'User successfully logs in and is redirected to dashboard',
            'priority': 'High'
        },
        {
            'feature': 'User Authentication',
            'test_id': 'TC002',
            'title': 'Verify login failure with invalid credentials',
            'steps': [
                'Navigate to login page',
                'Enter invalid username',
                'Enter invalid password',
                'Click Login button'
            ],
            'expected_result': 'Error message displayed and user remains on login page',
            'priority': 'High'
        },
        {
            'feature': 'User Profile',
            'test_id': 'TC003',
            'title': 'Verify user can update profile information',
            'steps': [
                'Login with valid credentials',
                'Navigate to profile page',
                'Update profile information',
                'Click Save button'
            ],
            'expected_result': 'Profile information is updated successfully with confirmation message',
            'priority': 'Medium'
        }
    ]


if __name__ == "__main__":
    # Test the export function with sample data
    sample_testcases = create_sample_testcases()
    save_testcases_to_excel(sample_testcases, "sample_testcases.xlsx")
    print("Sample Excel file created for testing purposes.")
